#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extrae las filas de España (y cualquier país) del Excel eRAIL a JSON.

Uso:
    python extraer_erail.py ES [DE ...]

Lee data/erail.xlsx (hoja Investigations y Safety recommendations)
→ data/erail/ES-investigations.json + data/erail/ES-recommendations.json
"""
import json
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
XLSX = RAIZ / "data" / "erail.xlsx"

NOMBRES_PAIS = {
    "ES": "Spain", "DE": "Germany", "FR": "France", "IT": "Italy", "PL": "Poland",
    "PT": "Portugal", "CZ": "Czech Republic", "RO": "Romania", "HU": "Hungary",
    "GB": "United Kingdom", "UK": "United Kingdom", "AT": "Austria", "DK": "Denmark",
    "FI": "Finland", "NO": "Norway", "SE": "Sweden", "BE": "Belgium", "IE": "Ireland",
    "NL": "The Netherlands", "LU": "Luxembourg", "EL": "Greece", "BG": "Bulgaria",
    "HR": "Croatia", "SI": "Slovenia", "SK": "Slovak Republic", "EE": "Estonia",
    "LV": "Latvia", "LT": "Lithuania", "CH": "Switzerland", "RS": "Serbia",
}


def serializar(v):
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10] if str(v)[:10] != "1900-01-01" else None
    return v


def extraer(codigo: str):
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    out_dir = RAIZ / "data" / "erail"
    out_dir.mkdir(parents=True, exist_ok=True)
    pais_excel = NOMBRES_PAIS.get(codigo, codigo)

    ws = wb["Investigations"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    i_pais = header.index("Country")
    i_erail = header.index("ERAIL Occurrence")
    filas = []
    for r in rows[1:]:
        if r[i_pais] == pais_excel:
            filas.append({h: serializar(v) for h, v in zip(header, r) if h})
    (out_dir / f"{codigo}-investigations.json").write_text(
        json.dumps(filas, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    print(f"[{codigo}] {len(filas)} investigaciones")

    ids = {f.get("ERAIL Occurrence") for f in filas}
    ws2 = wb["Safety recommendations"]
    rows2 = list(ws2.iter_rows(values_only=True))
    header2 = list(rows2[0])
    i_occ2 = header2.index("Occurrence")
    recs = []
    for r in rows2[1:]:
        if r[i_occ2] in ids:
            recs.append({h: serializar(v) for h, v in zip(header2, r) if h})
    (out_dir / f"{codigo}-recommendations.json").write_text(
        json.dumps(recs, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    print(f"[{codigo}] {len(recs)} recomendaciones vinculadas")
    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    for c in [a.upper() for a in sys.argv[1:]]:
        extraer(c)
